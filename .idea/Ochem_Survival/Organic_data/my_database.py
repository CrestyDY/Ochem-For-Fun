import openpyxl
import pandas as pd

dataframe = openpyxl.load_workbook("ochem.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

my_data = []

# Iterate the loop to read the cell values
for row in dataframe1.iter_rows(0, dataframe1.max_row):
    my_data.append([cell.value for cell in row])

new_data = []

for i in range(1, len(my_data)):
    if i==5820:
        continue
    new_list = []
    new_list.append(my_data[i][1])
    new_list.append(my_data[i][4])
    new_list.append(my_data[i][5])
    new_list.append(my_data[i][12])
    new_data.append(new_list)

data_without_duplicates = []
duplicate_dic = {}
for j in range(len(new_data)):
    if new_data[j][3] not in duplicate_dic.keys():
        data_without_duplicates.append(new_data[j])
        duplicate_dic[new_data[j][3]] = None
    else:
        continue

final_data = []
for i in range(len(data_without_duplicates)):
    if type(data_without_duplicates[i][1])==float or type(data_without_duplicates[i][1])==int:
        final_data.append(data_without_duplicates[i])
    else:
        continue

data_formula = []
for k in range(len(final_data)):
    data_formula.append(final_data[k][0])
max_len_formula = 0
for i in range(len(data_formula)):
    max_len_formula = max(max_len_formula, len(str(data_formula[i])))

data_pKa = []
for l in range(len(final_data)):
    data_pKa.append(final_data[l][1])
max_pKa = 0
for i in range(len(data_pKa)):
    max_pKa = max(max_pKa, len(str(data_pKa[i])))

data_temperature = []
for m in range(len(final_data)):
    data_temperature.append(final_data[m][2])
max_temperature = 0
for i in range(len(data_temperature)):
    max_temperature = max(max_temperature, len(str(data_temperature[i])))

data_iupac = []
for n in range(len(final_data)):
    data_iupac.append(final_data[n][3])
max_iupac = 0
for i in range(len(data_iupac)):
    max_iupac = max(max_iupac, len(str(data_iupac[i])))


"""print(data_without_duplicates,"\n\n\n\n")
print(len(final_data), len(data_without_duplicates), max_len_formula, max_pKa, max_temperature, max_iupac, "\n\n\n\n\n")
print(len(data_formula),"\n\n\n", len(data_temperature),"\n\n\n", len(data_iupac),"\n\n\n", len(data_pKa))"""
